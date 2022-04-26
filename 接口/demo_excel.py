"""
Excel 的基础用法，手工如何操作Excel，

手工操作excel的流程
1、打开Excel文件，（路径+文件名）
2、获取sheet表单
3、使用行号和列号 确定需要读取的数据
4、关闭文件

两个工具
openpyxl，支持xlsx 新型格式的读写，读取速度还可可以
tablib，支持多种格式读写， xlsx、xls、cvs、json、yaml、html，缺点读取速度慢
xlrd， 经典excel 读取库，局限性强，只能读取，不能写
pandas ，功能强大，太臃肿了，搞数据分析的时候可以用

安装openpyxl
pip install openpyxl
获取表单：
1、wb.active ,被选中，被激活
2、通过索引，wb.worksheets[索引]
3、通过sheet名字，wb['Sheet_name']


"""

import openpyxl

# 读取Excel文件
# 读取文件之前，一定要关闭文件
# windows 下面的路径有反斜杠，不想有转义的含义，前面加个r
wb = openpyxl.load_workbook(r'd:\cases.xlsx')
print(wb)

# active 是表示被激活的，被选择的sheet
active_sheet = wb.active
# sheetnames 列表当中存储的是字符串， _sheets里面存储的是对象


# 获取所有表单的正确用法
work_sheets = wb.worksheets
# 获取某一个表单，1、通过索引去获取
# sheet = wb.worksheets[0]
# print(sheet)


# 2、通过表单名称获取
# 过时，sheet.没有提示
# sheet = wb.get_sheet_by_name('Sheet1')
# print(sheet)
# 正确用法， sheet. 没有提示
sheet = wb['Sheet1']

# 读取单个单元格，行和列
# 行和列 是从1 开始的，不是python当中的 从0 开始的
# 可以从源码看出，第一个入参是行，第二个入参是列
cell = sheet.cell(1, 2)
print(cell)

# cell是一个对象，获取cell的值
print(cell.value)

# 获取某一行的数据
print(sheet[1])

# 获取某一行的值
for columm in sheet[1]:
    print(columm.value)

# 获取某一列
print(sheet['C'])

# 获取多行 1到3 行，第3行是包含的
print(sheet[1:3])
